from tkinter import*
import serial
import os
import threading 
from openpyxl import*



carpeta = r"C:\Users\prove\Documents\FERNY\PYTHON\Proyecto_Final"
ruta = r"C:\Users\prove\Documents\FERNY\PYTHON\Proyecto_Final\datosSensor.xlsx"
puerto= "COM3"

# Crear directorio o carpeta si no existe
if not os.path.exists(carpeta):
    os.makedirs(carpeta)

# Crear un archivo excel si no existe 
if not os.path. exists(ruta):
    libro = Workbook()
    libro.save(ruta)


# Crear puerto serial
try:
    arduino = serial.Serial(puerto, 9600)
except:
    print("No se estableció la conexión")


# ------------ FUNCIONES ---------------

def cerrar_ventana():
    arduino.close()
    raiz.destroy()

# BOTONES DE SALIDAS DIGITALES
def pressButton1():
    arduino.write(("1\n").encode())
    textoLabel1.config(text="LED 1 encendido", fg="blue")

def pressButton2():
    arduino.write(("2\n").encode())
    textoLabel1.config(text="LED 1 apagado" , fg="red")

def pressButton3():
    arduino.write(("3\n").encode())
    textoLabel2.config(text="LED 2 encendido", fg="blue")

def pressButton4():
    arduino.write(("4\n").encode())
    textoLabel2.config(text="LED 2 apagado" , fg="red")

def pressButton5():
    arduino.write(("5\n").encode())
    textoLabel3.config(text="LED 3 encendido", fg="blue")

def pressButton6():
    arduino.write(("6\n").encode())
    textoLabel3.config(text="LED 3 apagado" , fg="red")

def pressButton7():
    arduino.write(("7\n").encode())
    textoLabel4.config(text="LED 4 encendido", fg="blue")

def pressButton8():
    arduino.write(("8\n").encode())
    textoLabel4.config(text="LED 4 apagado" , fg="red")


# FX para leer y guardar datos en archivo de Excel
def read_and_save():
    # Cargar el archivo de Excel existente
    libro = load_workbook(ruta)

    #Selccionar la hoja activa
    hoja1 = libro.active
    hoja1.title = "Hoja_1"


    while arduino.is_open:
        #leer los datos del puerto serial
        data = arduino.readline().decode().strip()
      
        # Mostrar datos en la etiqueta label
        etiqueta1.config(text=data)

        #Obtener la siguiente fila vacía
        next_row = hoja1.max_row + 1 

        # Guardar datos en archivo de Excel
        hoja1.cell(row=next_row, column = 1).value = data 

        # Guardar el archivo de Excel
        libro.save(ruta)
        

# Crear un proceso paralelo a la ventana para leer y guardar los datos del Arduino
def start_thread():
    thread = threading.Thread(target= read_and_save)
    thread.daemon = True 
    thread.start()






#---------------------------------- VENTANA --------------------------------------

raiz = Tk()
raiz.title("Lectura de sensor y salidas de Arduino")
raiz.geometry("600x300")
raiz.protocol("WM_DELETE_WINDOW", cerrar_ventana)

raiz.columnconfigure(0, minsize= 200)

# - - - - -  SALIDAS DIGITALES  - - - - - 

#  BOTONES SALIDA 1 
button1= Button(raiz, text="ON", command=pressButton1, padx=10, pady=10, fg="green")
button1.grid(row=1, column=0)

button2= Button(raiz, text="OFF", command=pressButton2, padx=10, pady=10, fg="orange")
button2.grid(row=1, column=1)

textoLabel1 = Label(raiz, text="LED 1 apagado", fg="red", font=("Arial", 15))
textoLabel1.grid(row=1, column=2, padx=10, pady=10)

# BOTONES SALIDA 2 
button3= Button(raiz, text="ON", command=pressButton3, padx=10, pady=10, fg="green")
button3.grid(row=2, column=0)

button4= Button(raiz, text="OFF", command=pressButton4, padx=10, pady=10, fg="orange")
button4.grid(row=2, column=1)

textoLabel2 = Label(raiz, text="LED 2 apagado", fg="red", font=("Arial", 15))
textoLabel2.grid(row=2, column=2, padx=10, pady=10)

#  BOTONES SALIDA 3 
button5= Button(raiz, text="ON", command=pressButton5, padx=10, pady=10, fg="green")
button5.grid(row=3, column=0)

button6= Button(raiz, text="OFF", command=pressButton6, padx=10, pady=10, fg="orange")
button6.grid(row=3, column=1)

textoLabel3 = Label(raiz, text="LED 3 apagado", fg="red", font=("Arial", 15))
textoLabel3.grid(row=3, column=2, padx=10, pady=10)

#  BOTONES SALIDA 4 
button7= Button(raiz, text="ON", command=pressButton7, padx=10, pady=10, fg="green")
button7.grid(row=4, column=0)

button8= Button(raiz, text="OFF", command=pressButton8, padx=10, pady=10, fg="orange")
button8.grid(row=4, column=1)

textoLabel4 = Label(raiz, text="LED 4 apagado", fg="red", font=("Arial", 15))
textoLabel4.grid(row=4, column=2, padx=10, pady=10)

# - - - - -  DATOS DEL SENSOR  - - - - - -

# Botón para iniciar la lectura del sensor
boton_start = Button(raiz, text="Inicar", font=("Arial", 16), command=start_thread)
boton_start.grid(row=0, column=0, padx=10, pady=10)

#  Botón para detener y cerrar ventana
boton_stop = Button(raiz, text="Detener", font=("Arial", 16), command= cerrar_ventana)
boton_stop.grid(row=0, column=1, padx=10, pady=10)

# Label para mostrar los datos del sensor
etiqueta1= Label(raiz, text="Esperando datos", font=("Arial, 15"))
etiqueta1.grid(row=0, column=2, padx=10, pady=10)

# ------------------------
raiz.minsize(600,300)

raiz.mainloop()
